import logging
import os
import time

import gspread
import pandas as pd
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

logging.basicConfig(filename='./output/logfile.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s', level=logging.INFO)


class JobScrapper:
    """The class initialize a bot and searches for the job in LinkedIn and generates data in excel and google sheet.
    """
    def __init__(self, jobname, countryname, folder_id, editor_mail, cred_file_path) -> None:
        self.driver_path = './chromedriver/chromedriver.exe'
        self.jobname = jobname
        self.countryname = countryname
        self.folder_id = folder_id
        self.editor_mail = editor_mail
        self.cred_file_path = cred_file_path
        self.driver = webdriver.Chrome(self.driver_path)
        

    def get_url(self):
        """Creating the url based on the job role and job location.
        """
        
        job_url =self.jobname.replace(' ', '%20')
        country_url =self.countryname.replace(' ', '%20')

        url = f"https://www.linkedin.com/jobs/search?keywords={job_url}&location={country_url}&geoId=103644278&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0"
        logging.info(f'URL for the jobsearch:-  {url}')

        return url

    def get_browser(self):
        """Opening the browser and searching the url.
        """
        
        url = self.get_url()
        self.driver.get(url)
        logging.info('Browser is Opening')

    def get_total_job_no(self):
        """Get the no of total job search result.
        """
        jobs_num = self.driver.find_element(By.CSS_SELECTOR,"h1>span").get_attribute("innerText")

        if len(jobs_num.split(',')) > 1:
            jobs_num = int(jobs_num.split(',')[0])*1000
        else:
            jobs_num = int(jobs_num)

        jobs_num   = int(jobs_num)

        
        i = 2
        while i <= int(jobs_num/2)+1:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            i = i + 1
            print("Current at: ", i, "Percentage at: ", ((i+1)/(int(jobs_num/2)+1))*100, "%",end="\r")

            try:
                #We try to click on the load more results buttons in case it is already displayed.
                infinite_scroller_button = self.driver.find_element(By.XPATH, ".//button[@aria-label='See more jobs']")
                infinite_scroller_button.click()
                time.sleep(0.3)
            except:
                time.sleep(0.3)
                pass

    def job_basic_details(self):
        """Get the basic details of each job search result one by one.
        """
        job_lists = self.driver.find_element(By.CLASS_NAME,"jobs-search__results-list")
        jobs = job_lists.find_elements(By.TAG_NAME,"li") # return a list

        job_title_list = []
        company_name_list = []
        location_list = []
        date_list = []
        job_link_list = []

        for job in jobs:
            #job_title
            job_title = job.find_element(By.CSS_SELECTOR,"h3").get_attribute("innerText")
            job_title_list.append(job_title)
            
            #company_name
            company_name = job.find_element(By.CSS_SELECTOR,"h4").get_attribute("innerText")
            company_name_list.append(company_name)
            
            #location
            location = job.find_element(By.CSS_SELECTOR,"div>div>span").get_attribute("innerText")
            location_list.append(location)
            
            #date
            date = job.find_element(By.CSS_SELECTOR,"div>div>time").get_attribute("datetime")
            date_list.append(date)
            
            #job_link
            job_link = job.find_element(By.CSS_SELECTOR,"a").get_attribute("href")

            job_link_list.append(job_link)

        logging.info(f'job_link_list length, {len(job_link_list)}')
        df = pd.DataFrame({
                'Date': date,
                'Company': company_name_list,
                'Title': job_title_list,
                'Location': location_list,
                'Link': job_link_list
            })
        
        self.create_excel_files(df, job_link_list)
        self.make_google_sheets(df)
    
    def create_excel_files(self, df, job_link_list):
        """Create excel files populated with job search result.
        """
        df.to_excel('filenew.xlsx', index=False)

        workbook = load_workbook("filenew.xlsx")
        sheet = workbook.active
        row_count = sheet.max_row
        for i in range(2,row_count+1):
            sheet.cell(row=i, column=5).hyperlink = job_link_list[i-2]
            sheet.cell(row=i, column=5).style = "Hyperlink"
            workbook.save(f"./output/file-{self.countryname}-{self.jobname}.xlsx")

        os.remove("filenew.xlsx")

    def make_google_sheets(self, df):
        """Create Google sheet populated with job search result.
        """
        # Replace the values in the following lines with your own project ID and credentials file path
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = service_account.Credentials.from_service_account_file(self.cred_file_path, scopes = scope)

        service = build('drive', 'v3', credentials=creds)
        spreadsheet_name = f"file-{self.countryname}-{self.jobname}"

        query = "mimeType='application/vnd.google-apps.spreadsheet' and trashed=false and parents in '{}' and name='{}'".format(self.folder_id, spreadsheet_name)
        response = service.files().list(q=query, fields='nextPageToken, files(id, name)').execute()

        spreadsheet_id = None
        if 'files' in response:
            for file in response['files']:
                if file['name'] == spreadsheet_name:
                    spreadsheet_id = file['id']
                    break

        if spreadsheet_id:
            logging.info('Spreadsheet found in the folder')
        else:
            logging.info('Spreadsheet not found in the folder')

            file_metadata = {
                'name': spreadsheet_name,
                'parents': [self.folder_id],
                'mimeType': 'application/vnd.google-apps.spreadsheet'
            }
            spreadsheet = service.files().create(body=file_metadata).execute()
            spreadsheet_id = spreadsheet['id']

        self.give_editor_permission(service, spreadsheet_id)
        self.clear_spreadsheet(creds, spreadsheet_id)

        gc = gspread.authorize(creds)
        ss = gc.open_by_key(spreadsheet_id)
        worksheet = ss.get_worksheet(0)
        worksheet.update([df.columns.values.tolist()] + df.values.tolist())

        logging.info('making_google_sheet_ is_completed')

    def clear_spreadsheet(self, creds, spreadsheet_id):
        """Clear the spreadsheet.
        """
        sheet_service = build('sheets', 'v4', credentials=creds)
        spreadsheet = sheet_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_names = [sheet['properties']['title'] for sheet in spreadsheet['sheets']]
        sheet1 = sheet_names[0]

        range_name = f'{sheet1}!A1:Z10000'
        request = sheet_service.spreadsheets().values().clear(
                                spreadsheetId=spreadsheet_id,
                                range=range_name,
                                body={}
                            )
        response = request.execute()

    def give_editor_permission(self, service, spreadsheet_id):
        """Gives editor permission to a spreadsheet.
        """
        permission = {
                'type': 'user',
                'role': 'writer',
                'emailAddress': self.editor_mail
            }
        
        service.permissions().create(fileId=spreadsheet_id, body=permission).execute()
